from openpyxl import Workbook
from io import BytesIO, StringIO
from openpyxl.styles import Alignment
from pyexcel_io.manager import get_io
from pyexcel_io import save_data
class ExcelGraphBuilder:

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.x = 1
        self.y = 1
        self.row_max=0

    def add_table(self, data, title):

        self.ws.append([title])
        self.ws[f'A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

        self.row_max += 1

        for row in data:
            self.ws.append(row)
            self.row_max += 1
        col_count= len(data[0])
        cell = self.ws.cell(self.x, self.y)
        mcel = self.ws.cell(self.x, self.y + col_count)
        scell = cell.column_letter + str(cell.row)
        self.ws.merge_cells('%s:%s' % (
            scell,
            mcel.column_letter + str(mcel.row)
        ))

    def save(self, output=None):
        if output is None:
            output = BytesIO()
        self.wb.save(output)
        return output

    def save_ods(self, data,format_type="ods"):
        io= get_io(format_type)
        save_data(io, data,format_type)
        return io
